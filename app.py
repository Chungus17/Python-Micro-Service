from flask import Flask, jsonify, request
from flask_cors import CORS
import json
from urllib.request import Request, urlopen
import pandas as pd
import base64
from postmarker.core import PostmarkClient
import os
import threading
import requests
from collections import defaultdict
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime

app = Flask(__name__)
CORS(app)  

MY_API_KEY = os.environ.get("MY_API_KEY")
FM_TOKEN = os.environ.get("FM_TOKEN")
VERDI_API_KEY = os.environ.get("VERDI_API_KEY")
API_URL = "https://api.tookanapp.com/v2/get_fare_estimate"

POSTMARK_TOKEN = os.environ.get("POSTMARK_TOKEN")
senderEmail = "Support@tryverdi.com"
recipientEmail = "mrharoonkhan11@gmail.com"

def getData(start_date, end_date, filter_by, clientName, task_function):
    apiURL = f"https://tryverdi.com/api/transaction_data?user_id={filter_by}&start_date={start_date}&end_date={end_date}"

    headers = {
        "Authorization": f"Bearer {VERDI_API_KEY}"
    }

    response = requests.get(url=apiURL, headers=headers)

    if response.status_code == 200:
        try:
            data = response.json()
            if task_function == "hourly orders":
                get_Hourly_Orders(data, clientName)
            elif task_function == "average fare":
                get_Average_Fare(data, clientName)
            elif task_function == "number of orders":
                get_Number_Of_Orders(data, clientName)
            elif task_function == "total fare":
                get_Total_Fare(data, clientName)
            elif task_function == "amount ranges":
                get_Amount_Ranges(data, clientName)
            elif task_function == "pickup counts":
                get_Pickup_Counts_Per_Area(data, clientName)
            else:
                return jsonify({
                    "message": "Wrong function called 😂"
                })
        except ValueError as e:
            return("Failed to parse JSON. Raw response was:")
    else:
        return(f"Request failed with status code {response.status_code}")

def send_email(excel_file, subject, clientName):
    try:
        
        # Read the file and encode in base64
        with open(excel_file, "rb") as f:
            excel_data = f.read()
            encoded_excel = base64.b64encode(excel_data).decode()

        # Send email using Postmark
        client = PostmarkClient(server_token=POSTMARK_TOKEN)
        client.emails.send(
            From=senderEmail,
            To=recipientEmail,
            Subject=subject,
            HtmlBody="<strong>Please find the attached Excel file.</strong>",
            Attachments=[
                {
                    "Name": f"{clientName} {excel_file}",
                    "Content": encoded_excel,
                    "ContentType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                }
            ]
        )

    except Exception as e:
        print(f"Error sending email: {str(e)}")


# !-----------------------------------------------------------------------------------------------------------------------------------------------------------------
#* ---------------------------------------------------DATA ANALYSIS ENDPOINT----------------------------------------------------------------------------------------

# TODO: Endpoint for getting NUMBER OF ORDERS PER HOUR (Specific client or all clients)
@app.route('/data_analysis', methods=["POST"])
def data_analysis():

    params_data = request.get_json()
    start_date = params_data.get('start_date')
    end_date = params_data.get('end_date')
    filter_by = params_data.get('filter_by')
    clientName = params_data.get('clientName')
    task_function = params_data.get('task_function')

    thread = threading.Thread(target=getData, args=(start_date, end_date, filter_by, clientName, task_function))
    thread.start()

    return jsonify({
        "message": "You will receive the report in your email inbox in less than a minute!"
    })


# !-----------------------------------------------------------------------------------------------------------------------------------------------------------------
#* -----------------------------------------------ALL FUNCTIONS FOR DATA ANALYSIS-----------------------------------------------------------------------------------
def get_Hourly_Orders(data, clientName):
    # Step 1: Prepare a nested dictionary {user_name: {hour: count}}
    order_counts = defaultdict(lambda: defaultdict(int))

    for order in data:
        user = order['user_name']
        created_at = datetime.strptime(order['created_at'], "%Y-%m-%d %H:%M:%S")
        hour = created_at.hour  # Extract the hour (0–23)
        order_counts[user][hour] += 1

    # Step 2: Format into a list of dictionaries for DataFrame
    FinalData = []
    for user, hourly_counts in order_counts.items():
        row = {'user_name': user}
        total = 0
        for hour in range(24):
            count = hourly_counts.get(hour, 0)
            row[f'{hour:02d}:00-{(hour+1)%24:02d}:00'] = count
            total += count
        row['Total'] = total
        FinalData.append(row)

    # Step 3: Create DataFrame
    df = pd.DataFrame(FinalData)

    # Step 4: Optional - sort columns by hour
    hour_columns = [f'{hour:02d}:00-{(hour+1)%24:02d}:00' for hour in range(24)]
    df = df[['user_name'] + hour_columns + ['Total']]

    # Step 5: Save to Excel
    excel_file = f"{clientName}_hourly_order_report.xlsx"
    df.to_excel(excel_file, index=False)

    # Load workbook to adjust column widths
    workbook = load_workbook(excel_file)
    sheet = workbook.active

    # Set width of first column ('user_name') to 16
    sheet.column_dimensions[get_column_letter(1)].width = 23

    # Set width of remaining columns (hourly columns + 'Total') to 13
    for col in range(2, sheet.max_column + 1):
        sheet.column_dimensions[get_column_letter(col)].width = 13

    # Save the workbook
    workbook.save(excel_file)

    send_email(excel_file, subject="Hourly orders", clientName=clientName)

def get_Average_Fare(data, clientName):
    output_filename = f"Average Fare.xlsx"

    # Step 1: Group absolute amounts by user_name
    user_amounts = defaultdict(list)

    for entry in data:
            user = entry.get('user_name')
            amount_str = entry.get('amount', '0')
            try:
                amount = abs(float(amount_str))
                user_amounts[user].append(amount)
            except ValueError:
                print(f"Skipping invalid amount for user {user}: {amount_str}")
                continue

        # Step 2: Compute averages
    final_data = []
    for user, amounts in user_amounts.items():
            total = sum(amounts)
            avg = sum(amounts) / len(amounts) if amounts else 0
            final_data.append({
                "User Name": user,
                "Average Fare": round(avg, 2)
            })


        # Step 3: Create DataFrame
    df = pd.DataFrame(final_data)

        # Step 4: Save to Excel
    df.to_excel(output_filename, index=False)

        # Step 5: Adjust column widths
    workbook = load_workbook(output_filename)
    sheet = workbook.active

    sheet.column_dimensions[get_column_letter(1)].width = 25  # User_Name
    sheet.column_dimensions[get_column_letter(2)].width = 18  # Average_Amount

    workbook.save(output_filename)

    send_email(output_filename, subject="Average Fare", clientName=clientName)
    
def get_Number_Of_Orders(data, clientName):

    output_file_name = "Total number of orders.xlsx"
    # Create a DataFrame
    df = pd.DataFrame(data)

    # Count the number of orders per user_name
    order_counts = df['user_name'].value_counts().reset_index()
    order_counts.columns = ['User Name', 'Number of Orders']

    # Export to Excel
    order_counts.to_excel(output_file_name, index=False)

        # Step 5: Adjust column widths
    workbook = load_workbook(output_file_name)
    sheet = workbook.active

    sheet.column_dimensions[get_column_letter(1)].width = 25  # User_Name
    sheet.column_dimensions[get_column_letter(2)].width = 15  # Average_Amount

    workbook.save(output_file_name)

    send_email(output_file_name, subject="Number of orders", clientName=clientName)

def get_Total_Fare(data, clientName):
    output_filename = f"Total Fare.xlsx"

    # Step 1: Group absolute amounts by user_name
    user_amounts = defaultdict(list)

    for entry in data:
            user = entry.get('user_name')
            amount_str = entry.get('amount', '0')
            try:
                amount = abs(float(amount_str))
                user_amounts[user].append(amount)
            except ValueError:
                print(f"Skipping invalid amount for user {user}: {amount_str}")
                continue

        # Step 2: Compute averages
    final_data = []
    for user, amounts in user_amounts.items():
            total = sum(amounts) if amounts else 0
            final_data.append({
                "User Name": user,
                "Total Fare": round(total, 2)
            })


    # Step 3: Create DataFrame
    df = pd.DataFrame(final_data)

    # Step 4: Save to Excel
    df.to_excel(output_filename, index=False)

    # Step 5: Adjust column widths
    workbook = load_workbook(output_filename)
    sheet = workbook.active

    sheet.column_dimensions[get_column_letter(1)].width = 25  # User_Name
    sheet.column_dimensions[get_column_letter(2)].width = 18  # Average_Amount

    workbook.save(output_filename)

    send_email(output_filename, subject="Total Fare", clientName=clientName)

def get_Amount_Ranges(data, clientName):

    output_file_name = "Orders in fare ranges.xlsx"

    # Define range intervals
    step = 0.25
    range_limits = [(round(start, 2), round(start + step, 2)) for start in [1 + i * step for i in range(8)]]  
    range_labels = [f"{lo}-{hi}" for lo, hi in range_limits]

    # Prepare data storage
    user_range_counts = defaultdict(lambda: defaultdict(int))

    # Process each record
    for record in data:
        user = record['user_name']
        amount = abs(float(record['amount']))
        for lo, hi in range_limits:
            if lo <= amount < hi:
                label = f"{lo}-{hi}"
                user_range_counts[user][label] += 1
                break

    # Convert to DataFrame
    output_rows = []
    for user, counts in user_range_counts.items():
        row = {'user_name': user}
        for label in range_labels:
            row[label] = counts.get(label, 0)
        output_rows.append(row)

    df = pd.DataFrame(output_rows)
    df.to_excel(output_file_name, index=False)
    # Step 5: Adjust column widths
    workbook = load_workbook(output_file_name)    
    sheet = workbook.active

    sheet.column_dimensions[get_column_letter(1)].width = 25  # User_Name
    # Save the updated workbook
    workbook.save(output_file_name)

    send_email(output_file_name, subject="Total orders in fare ranges", clientName=clientName)

def get_Pickup_Counts_Per_Area(data, clientName):

    output_filename = "Pickups per area per hour.xlsx"

    # Load JSON data from file
    with open('areas.json', 'r', encoding='utf-8') as file:
        neighborhood_data = json.load(file)

    # Build a mapping of each alias → canonical name
    area_alias_map = {}
    for item in neighborhood_data:
        if 'neighborhoodenglish' in item:
            aliases = [alias.strip() for alias in item['neighborhoodenglish'].split(',')]
            if aliases:
                canonical = aliases[0]
                for alias in aliases:
                    area_alias_map[alias.lower()] = canonical

    # Area extraction function
    def extract_area_simple(address, alias_map):
        if not address:
            return "Unknown"
        address_lower = address.lower()
        for alias in alias_map:
            if alias in address_lower:
                return alias_map[alias]
        return "Unknown"

    # Enrich each object with extracted areas
    for item in data:
        pickup_address = item.get('pickup_task', {}).get('address')
        delivery_address = item.get('delivery_task', {}).get('address')

        item['pickup_area'] = extract_area_simple(pickup_address, area_alias_map)
        item['delivery_area'] = extract_area_simple(delivery_address, area_alias_map)

    # Initialize a nested dictionary to store area → hour → count
    area_hour_count = {}

    for item in data:
        area = item.get('pickup_area', 'Unknown')
        timestamp = item.get('created_at', '')

        try:
            dt = datetime.strptime(timestamp, "%Y-%m-%d %H:%M:%S")
            hour = dt.hour
        except:
            continue

        if area not in area_hour_count:
            area_hour_count[area] = [0] * 24

        area_hour_count[area][hour] += 1

    # Prepare rows and sort by area name
    rows = []
    for area, hours in sorted(area_hour_count.items()):
        total = sum(hours)
        row = [area] + hours + [total]
        rows.append(row)

    columns = ['Pickup Area'] + [f"{i}-{i+1}" for i in range(24)] + ['Total']

    # Create DataFrame and save to Excel
    df = pd.DataFrame(rows, columns=columns)
    excel_path = output_filename
    df.to_excel(excel_path, index=False)

    # Adjust the first column width
    wb = load_workbook(excel_path)
    ws = wb.active
    ws.column_dimensions[get_column_letter(1)].width = 30  # Wider first column
    wb.save(excel_path)

    send_email(output_filename, subject="Pickups per area in every hour", clientName=clientName)


# !-----------------------------------------------------------------------------------------------------------------------------------------------------------------


if __name__ == '__main__':
    app.run(debug=False)
